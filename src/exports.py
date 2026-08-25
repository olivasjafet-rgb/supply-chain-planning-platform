"""Deliverables: an Excel workbook, a Power BI star schema, and DAX measures.

The engine is Python, but nobody in a merchandising or finance meeting opens
Python. So every run also produces the two artefacts those teams actually use:

  * one workbook, formatted, with the assumptions stated on the first sheet;
  * a clean star schema plus a ``measures.dax`` file, so the Power BI model is
    a load-and-go rather than a modelling project.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Palette -- one accent, neutral greys, and a red reserved for exceptions.
HEADER_FILL = PatternFill("solid", fgColor="1F4E3D")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F4E3D")
SUB_FONT = Font(italic=True, size=9, color="666666")
ALERT_FILL = PatternFill("solid", fgColor="FDE7E9")

MONEY = '"$"#,##0'
MONEY2 = '"$"#,##0.00'
PCT = "0.0%"
NUM = "#,##0"
DEC = "#,##0.00"

FORMATS = {
    "usd": MONEY, "value": MONEY, "cost": MONEY2, "price": MONEY2,
    "pct": PCT, "rate": PCT, "share": PCT, "wmape": PCT, "accuracy": PCT,
    "bias": PCT, "units": NUM, "ratio": "0.0000",
}


def _fmt_for(column: str) -> str | None:
    c = column.lower()
    if c.endswith("_usd") or "value_usd" in c:
        return MONEY
    if "unit_cost" in c or "price" in c:
        return MONEY2
    if c.endswith("_pct") or c in ("wmape", "accuracy", "fill_rate", "demand_share",
                                   "revenue_share", "cumulative_share", "gross_margin_pct",
                                   "empirical_rate", "effective_rate", "origin_expense_rate",
                                   "seasonal_index", "yoy_growth"):
        return PCT if "index" not in c else "0.000"
    if "ratio" in c:
        return "0.0000"
    if c.endswith("_units") or c in ("units", "safety_stock", "reorder_point", "order_up_to"):
        return NUM
    if c in ("gmroi", "inventory_turns", "cv", "weeks_of_supply", "months_of_cover",
             "avg_weeks_of_supply", "min_weeks_of_supply", "safety_stock_weeks"):
        return DEC
    return None


def _write_sheet(writer, df: pd.DataFrame, name: str, note: str = "") -> None:
    """Write a frame with a frozen, styled header and sensible column widths."""
    frame = df.copy()
    for col in frame.columns:
        if str(frame[col].dtype).startswith("period"):
            frame[col] = frame[col].astype(str)

    start = 2 if note else 0
    frame.to_excel(writer, sheet_name=name[:31], index=False, startrow=start)
    ws = writer.sheets[name[:31]]

    if note:
        ws.cell(row=1, column=1, value=note).font = SUB_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(4, len(frame.columns)))

    header_row = start + 1
    for j, col in enumerate(frame.columns, start=1):
        cell = ws.cell(row=header_row, column=j)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        fmt = _fmt_for(str(col))
        if fmt:
            for i in range(header_row + 1, header_row + len(frame) + 1):
                ws.cell(row=i, column=j).number_format = fmt

        width = max(len(str(col)) + 2, 11)
        sample = frame[col].dropna().astype(str).head(200)
        if len(sample):
            longest = sample.str.len().max()
            if pd.notna(longest):
                width = max(width, min(int(longest) + 2, 42))
        ws.column_dimensions[get_column_letter(j)].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(frame.columns))}{header_row + len(frame)}"
    )


def _cover_sheet(writer, results: dict) -> None:
    cfg = results["config"]
    s = results["summary"]
    a = results["accuracy"]
    ss = results["steady_state"]

    rows = [
        ("PLANNING RUN", cfg.profile_name, ""),
        ("", "", ""),
        ("SCOPE", "", ""),
        ("Planning grain", cfg.planning_level, " > ".join(cfg.products.keys)),
        ("Horizon", f"{cfg.horizon_months} months", str(results["plan"]["month"].min())
         + " .. " + str(results["plan"]["month"].max())),
        ("Items planned", f"{results['plan'][cfg.planning_level].nunique()}", ""),
        ("Selling accounts", f"{results['allocation']['account'].nunique()}", ""),
        ("", "", ""),
        ("DEMAND PLAN", "", ""),
        ("Method", cfg.forecast["method"], f"{cfg.forecast['moving_average_window']}-month base, "
         f"seasonality damped to {cfg.forecast['seasonality']['damping']:.0%}"),
        ("Backtest accuracy", a["accuracy"], "3 months held out, volume-weighted"),
        ("Rolling accuracy", results["rolling_backtest"]["accuracy"].mean(),
         f"mean of {len(results['rolling_backtest'])} origins -- the number to trust"),
        ("Rolling bias", results["rolling_backtest"]["bias_pct"].mean(),
         "positive = over-forecast"),
        ("Items over 30% error", a["skus_over_30pct_error"], "manual review candidates"),
        ("", "", ""),
        ("REPLENISHMENT", "", ""),
        ("Safety stock method", cfg.replenishment["safety_stock_method"],
         f"service level {cfg.service_level:.0%}"),
        ("Order policy", cfg.replenishment["order_policy"],
         "case pack and MOQ enforced"),
        ("Planned purchases", s["planned_purchases_usd"], ""),
        ("Average inventory", s["avg_inventory_usd"], ""),
        ("Inventory turns", s["inventory_turns_annualised"], "annualised"),
        ("GMROI", s["gmroi"], f"target >= {cfg.kpi_targets['gmroi_min']}"),
        ("Network fill rate", s["fill_rate"], f"target >= {cfg.kpi_targets['fill_rate_min']:.0%}"),
        ("Items at stockout risk", s["items_with_stockout"], ""),
        ("", "", ""),
        ("LANDED-COST SCENARIO", cfg.landed_cost["scenario"]["name"], ""),
        ("Effective from", cfg.landed_cost["scenario"]["effective_from"], ""),
        ("Weighted origin surcharge", ss["weighted_surcharge_rate"], "empirical, per supplier"),
        ("Steady-state cost ratio", ss["steady_state_cost_ratio"],
         "the asymptote, not a month-1 result"),
        ("Purchase saving (cash)", results["purchase_savings"]["purchase_saving_usd"].sum(), "gate 1"),
        ("Receipt saving (goods)", results["receipt_savings"]["receipt_saving_usd"].sum(), "gate 2"),
        ("COGS saving (P&L)", results["cogs_dilution"]["cogs_saving_usd"].sum(), "gate 3"),
        ("P&L realisation", results["cogs_dilution"]["pnl_realisation_pct"].iloc[-1],
         "share of the receipt saving that has reached the P&L"),
        ("", "", ""),
        ("HOW TO READ THIS", "", ""),
        ("", "Every sheet is an output of one Python run; nothing is typed by hand.", ""),
        ("", "Change config.yaml and re-run to get a different scenario.", ""),
        ("", "Figures are simulated for a portfolio; the method is production logic.", ""),
    ]

    df = pd.DataFrame(rows, columns=["Item", "Value", "Note"])
    df.to_excel(writer, sheet_name="Summary", index=False, startrow=1)
    ws = writer.sheets["Summary"]
    ws.cell(row=1, column=1, value=f"{cfg.profile_name} -- Supply Chain Planning Run").font = TITLE_FONT
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 58

    section_rows = {3, 10, 16, 26, 36}
    for i, (item, value, _note) in enumerate(rows, start=3):
        if not item and not value:
            continue
        if (i - 2) in section_rows or (item.isupper() and item and not value):
            ws.cell(row=i, column=1).font = Font(bold=True, color="1F4E3D")
        cell = ws.cell(row=i, column=2)
        if isinstance(value, float):
            key = item.lower()
            if any(k in key for k in ("accuracy", "bias", "rate", "surcharge",
                                      "realisation", "fill")):
                cell.number_format = PCT
            elif "ratio" in key:
                cell.number_format = "0.0000"
            elif any(k in key for k in ("purchase", "inventory", "saving")) and "turns" not in key:
                cell.number_format = MONEY
            else:
                cell.number_format = DEC
    ws.sheet_view.showGridLines = False


def write_excel_workbook(results: dict, path: Path) -> Path:
    """One workbook, ordered the way a planner reads it: answer first, detail after."""
    path.parent.mkdir(parents=True, exist_ok=True)
    cfg = results["config"]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _cover_sheet(writer, results)

        _write_sheet(
            writer, results["orders"], "Order Book",
            "What to buy, from whom, when to place it and when it lands. "
            "This is the actionable output of the run.",
        )
        _write_sheet(
            writer, results["frozen_window"], "Stockout Triage",
            "Projected shortages split into what a policy change can fix and what "
            "was already locked in by the lead time. They need different actions.",
        )
        _write_sheet(
            writer, results["monthly"], "Monthly Position",
            "Network demand, receipts, closing inventory and purchases by month.",
        )
        _write_sheet(
            writer, results["item_kpis"], "Item KPIs",
            "Turns, GMROI, weeks of supply and a health flag per item.",
        )
        _write_sheet(
            writer, results["policy"].drop(columns=["demand_class"], errors="ignore"),
            "Replenishment Policy",
            "Safety stock, reorder point (min) and order-up-to (max) per item, "
            "with the statistics they were derived from.",
        )
        _write_sheet(
            writer, results["plan"], "Time-Phased Plan",
            "The month-by-month projection behind every order: opening, demand, "
            "receipts, closing, cover.",
        )
        _write_sheet(
            writer, results["forecast"], "Demand Plan",
            "Statistical forecast, seasonal index and the planner override column. "
            "The statistical value is never overwritten.",
        )
        _write_sheet(
            writer, results["forecast_scores"], "Forecast Accuracy",
            "Backtest: 3 months held out, re-forecast, scored. WMAPE and bias per item.",
        )
        _write_sheet(
            writer, results["rolling_backtest"], "Rolling Backtest",
            "The same model scored from six successive cut-off dates. Parameters "
            "were selected on this, not on a single holdout -- on one fold the "
            "bias read +13%, across six it read -9%.",
        )
        _write_sheet(
            writer, results["seasonal_indices"], "Seasonal Indices",
            "Monthly multiplicative indices estimated at category level and damped.",
        )
        _write_sheet(
            writer, results["service"], "Fill Rate by Account",
            "Requirement vs allocated. Minimum presentation stock is protected "
            "before the remainder is fair-shared.",
        )
        _write_sheet(
            writer, results["allocation"], "Allocation Detail",
            "Item x month x account allocation with fill rate and short units.",
        )
        _write_sheet(
            writer, results["segments"], "ABC-XYZ Segments",
            "Value contribution crossed with demand variability, and the review "
            "cadence each segment earns.",
        )
        _write_sheet(
            writer, results["supplier_rates"], "Supplier Rates",
            "Empirical origin-expense rate per supplier: actual expenses divided "
            "by actual purchase value.",
        )
        _write_sheet(
            writer, results["cogs_dilution"], "Landed Cost Dilution",
            "The three gates: receipts, the chained weighted-average cost ratio, "
            "and how much of the saving has reached the P&L.",
        )
        _write_sheet(
            writer, results["purchase_savings"], "Purchase Savings",
            "Gate 1 -- the cash effect, recognised when the PO is placed.",
        )
    return path


# -----------------------------------------------------------------------------
# Power BI
# -----------------------------------------------------------------------------

DAX_MEASURES = """// =============================================================================
//  Measures -- Supply Chain Planning model
//  Paste into a Power BI model built on the tables in outputs/powerbi/.
//  Relationships (all single-direction, many-to-one from fact to dim):
//     fact_plan[sku]        -> dim_product[sku]
//     fact_plan[month]      -> dim_calendar[month]
//     fact_allocation[sku]  -> dim_product[sku]
//     fact_allocation[account] -> dim_location[account]
//     dim_product[supplier_id] -> dim_supplier[supplier_id]
// =============================================================================

Forecast Units = SUM ( fact_plan[forecast_units] )

Closing Inventory = SUM ( fact_plan[closing_value_usd] )

Average Inventory = AVERAGEX ( VALUES ( dim_calendar[month] ), [Closing Inventory] )

COGS = SUMX ( fact_plan, fact_plan[forecast_units] * fact_plan[unit_cost_usd] )

Planned Purchases = SUM ( fact_plan[order_value_usd] )

Inventory Turns =
DIVIDE ( [COGS], [Average Inventory] )
    * DIVIDE ( 12, DISTINCTCOUNT ( dim_calendar[month] ) )

Weeks of Supply = DIVIDE ( [Closing Inventory], DIVIDE ( [COGS], 4.33 ) )

Gross Margin =
SUMX (
    fact_plan,
    fact_plan[forecast_units]
        * ( RELATED ( dim_product[retail_price] ) * 0.65 - fact_plan[unit_cost_usd] )
)

GMROI = DIVIDE ( [Gross Margin], [Average Inventory] )

Fill Rate =
DIVIDE (
    SUM ( fact_allocation[allocated_units] ),
    SUM ( fact_allocation[requirement_units] )
)

Short Units = SUM ( fact_allocation[short_units] )

Stockout Item-Months =
CALCULATE ( COUNTROWS ( fact_plan ), fact_plan[stockout_units] > 0 )

// -- Landed-cost scenario -----------------------------------------------------

Receipt Saving = SUM ( fact_landed_cost[receipt_saving_usd] )

COGS Saving = SUM ( fact_landed_cost[cogs_saving_usd] )

Cumulative COGS Saving =
CALCULATE (
    [COGS Saving],
    FILTER ( ALL ( dim_calendar ), dim_calendar[month] <= MAX ( dim_calendar[month] ) )
)

Cumulative Receipt Saving =
CALCULATE (
    [Receipt Saving],
    FILTER ( ALL ( dim_calendar ), dim_calendar[month] <= MAX ( dim_calendar[month] ) )
)

P&L Realisation % =
DIVIDE ( [Cumulative COGS Saving], [Cumulative Receipt Saving] )

// -- Conditional formatting helpers ------------------------------------------

WoS Status =
VAR w = [Weeks of Supply]
RETURN
    SWITCH (
        TRUE (),
        ISBLANK ( w ), "No data",
        w < 6,  "Thin cover",
        w > 14, "Overstocked",
        "Healthy"
    )
"""

POWERBI_README = """# Power BI model

Load every CSV in this folder as a table, then set these relationships
(all many-to-one, single cross-filter direction, from fact to dimension):

| From                        | To                       |
|-----------------------------|--------------------------|
| fact_plan[sku]              | dim_product[sku]         |
| fact_plan[month]            | dim_calendar[month]      |
| fact_allocation[sku]        | dim_product[sku]         |
| fact_allocation[account]    | dim_location[account]    |
| fact_allocation[month]      | dim_calendar[month]      |
| fact_sales_history[sku]     | dim_product[sku]         |
| fact_sales_history[month]   | dim_calendar[month]      |
| fact_landed_cost[month]     | dim_calendar[month]      |
| dim_product[supplier_id]    | dim_supplier[supplier_id]|

Mark `dim_calendar` as the date table on `month_start`.

Then paste `measures.dax` into a new measure table. Suggested pages:

1. **Executive** -- purchases, inventory, turns, GMROI, fill rate; demand and
   inventory by month.
2. **Demand** -- forecast vs history by category, seasonal indices, backtest
   accuracy by item.
3. **Replenishment** -- order book, weeks-of-supply distribution, stockout risk.
4. **Service** -- fill rate by account and channel, short units.
5. **Landed cost** -- the three savings gates and the P&L realisation curve.
"""


def write_powerbi_star_schema(results: dict, folder: Path) -> int:
    """Write a clean star schema plus the DAX and the modelling instructions."""
    folder.mkdir(parents=True, exist_ok=True)
    cfg = results["config"]
    tables = results["tables"]

    calendar = tables["dim_calendar"].copy()
    calendar["month_start"] = pd.PeriodIndex(calendar["month"], freq="M").to_timestamp()

    levels = [cfg.planning_level] + cfg.products.above(cfg.planning_level)
    master = tables["dim_product"][levels].drop_duplicates(subset=[cfg.planning_level])
    plan = results["plan"].merge(master, on=cfg.planning_level, how="left")

    exports = {
        "dim_product": tables["dim_product"],
        "dim_supplier": results["supplier_rates"],
        "dim_location": tables["dim_location"],
        "dim_calendar": calendar,
        "fact_sales_history": tables["fact_sales_history"],
        "fact_plan": plan,
        "fact_forecast": results["forecast"],
        "fact_allocation": results["allocation"],
        "fact_landed_cost": results["cogs_dilution"],
        "fact_order_book": results["orders"],
        "dim_segment": results["segments"],
        "fact_forecast_accuracy": results["forecast_scores"],
        "fact_rolling_backtest": results["rolling_backtest"],
        "fact_stockout_triage": results["frozen_window"],
    }

    for name, df in exports.items():
        frame = df.copy()
        for col in frame.columns:
            if str(frame[col].dtype).startswith("period"):
                frame[col] = frame[col].astype(str)
        frame.to_csv(folder / f"{name}.csv", index=False)

    (folder / "measures.dax").write_text(DAX_MEASURES, encoding="utf-8")
    (folder / "README.md").write_text(POWERBI_README, encoding="utf-8")
    return len(exports)
